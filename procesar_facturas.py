from pathlib import Path
import tkinter as tk
from tkinter import filedialog
import xml.etree.ElementTree as ET
from datetime import datetime
import openpyxl
from openpyxl.styles import Font


# ==========================================
# Namespaces del CFDI 4.0
# ==========================================

ns = {
    "cfdi": "http://www.sat.gob.mx/cfd/4",
    "tfd": "http://www.sat.gob.mx/TimbreFiscalDigital"
}

MESES_ES = {
    1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
    7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic",
}


# ==========================================
# Catálogo de conceptos conocidos
# ------------------------------------------
# La Descripcion del XML se parte en (CONCEPTO, COMPLEMENTO) buscando cuál
# de estas categorías aparece al inicio. Ordenado del más largo al más
# corto para que gane el match más específico (p. ej. "MANTENIMIENTO
# PREVENTIVO DE UNIDADES" antes que "MANTENIMIENTO PREVENTIVO").
#
# Si aparece una descripción que no calza con ninguna, el script la deja
# completa en CONCEPTO, avisa en consola, y tú agregas la categoría nueva
# aquí para la próxima corrida.
# ==========================================

CATALOGO_CONCEPTOS = [
    "SERVICIO DE LIMPIEZA Y DESINFECCION",
    "REPARACION DE CAJA SECA",
    "MANTENIMIENTO PREVENTIVO",
    "MONITOREO A DISTANCIA",
    "ESTUDIOS SOCIOECONOMICO",
    "CAMBIO DE PUERTAS",
    "CAMBIO DE PISO",
    "LLANTAS",
]
CATALOGO_CONCEPTOS.sort(key=len, reverse=True)

ENCABEZADOS = [
    "FOLIO", "FECHA", "RFC EMISOR", "EMISOR", "RFC CLIENTE", "CLIENTE",
    "CONCEPTO", "COMPLEMENTO", "PERIODO", "IMPORTE", "IVA", "TOTAL", "UUID",
]


# ==========================================
# Helpers
# ==========================================

def separar_concepto(descripcion):
    """Separa la Descripcion del XML en (CONCEPTO, COMPLEMENTO) usando
    CATALOGO_CONCEPTOS. Si no reconoce ninguna categoría, regresa toda la
    descripción como CONCEPTO y COMPLEMENTO vacío."""

    desc = " ".join(descripcion.split())  # normaliza espacios/saltos

    for categoria in CATALOGO_CONCEPTOS:
        if desc.upper().startswith(categoria):
            resto = desc[len(categoria):].strip()
            if resto.upper().startswith("DE "):
                resto = resto[3:].strip()
            resto = resto.lstrip(",").strip()
            return categoria, resto

    print(f"  AVISO: descripción no reconocida en el catálogo: {desc!r}")
    return desc, ""


def periodo_desde_fecha(fecha_iso):
    fecha = datetime.fromisoformat(fecha_iso)
    return f"{MESES_ES[fecha.month]}-{str(fecha.year)[2:]}"


def fecha_desde_iso(fecha_iso):
    return datetime.fromisoformat(fecha_iso).strftime("%d/%m/%Y")


def extraer_filas(xml_path):
    """Regresa una lista de filas (una por concepto) para un XML dado."""

    tree = ET.parse(xml_path)
    root_xml = tree.getroot()

    emisor = root_xml.find("cfdi:Emisor", ns)
    receptor = root_xml.find("cfdi:Receptor", ns)
    timbre = root_xml.find(".//tfd:TimbreFiscalDigital", ns)

    emisor_nombre = emisor.attrib.get("Nombre") if emisor is not None else None
    emisor_rfc = emisor.attrib.get("Rfc") if emisor is not None else None
    receptor_rfc = receptor.attrib.get("Rfc") if receptor is not None else None
    receptor_nombre = receptor.attrib.get("Nombre") if receptor is not None else None
    uuid = timbre.attrib.get("UUID") if timbre is not None else None

    folio = root_xml.attrib.get("Folio")
    fecha_iso = root_xml.attrib.get("Fecha")
    fecha = fecha_desde_iso(fecha_iso)
    periodo = periodo_desde_fecha(fecha_iso)

    filas = []

    conceptos = root_xml.find("cfdi:Conceptos", ns)
    if conceptos is None:
        return filas

    for concepto in conceptos.findall("cfdi:Concepto", ns):
        descripcion = concepto.attrib.get("Descripcion", "")
        importe = float(concepto.attrib.get("Importe", 0))

        traslado = concepto.find(
            ".//cfdi:Impuestos/cfdi:Traslados/cfdi:Traslado", ns
        )
        iva = float(traslado.attrib.get("Importe", 0)) if traslado is not None else 0.0

        total = round(importe + iva, 2)

        concepto_cat, complemento = separar_concepto(descripcion)

        filas.append([
            folio, fecha, emisor_rfc, emisor_nombre, receptor_rfc, receptor_nombre,
            concepto_cat, complemento, periodo,
            importe, iva, total, uuid,
        ])

    return filas


# ==========================================
# Programa principal
# ==========================================

def main():
    root = tk.Tk()
    root.withdraw()

    directorio = filedialog.askdirectory(
        title="Selecciona el directorio que contiene los XML"
    )
    if not directorio:
        print("No seleccionaste ningún directorio.")
        return
    directorio = Path(directorio)

    archivo_salida = filedialog.asksaveasfilename(
        title="Selecciona (o crea) el Excel donde se va guardando la información",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile="facturas_procesadas.xlsx",
    )
    if not archivo_salida:
        print("No seleccionaste ningún archivo de salida.")
        return
    archivo_salida = Path(archivo_salida)

    archivos_xml = list(directorio.glob("*.xml"))
    print(f"\nXML encontrados: {len(archivos_xml)}")
    if not archivos_xml:
        return

    # ------------------------------------------
    # Abrir el Excel existente (para ir agregando) o crear uno nuevo
    # ------------------------------------------

    if archivo_salida.exists():
        wb = openpyxl.load_workbook(archivo_salida)
        ws = wb["Facturas"] if "Facturas" in wb.sheetnames else wb.active
        col_uuid = ENCABEZADOS.index("UUID")
        uuids_existentes = {
            row[col_uuid].value
            for row in ws.iter_rows(min_row=2)
            if row[col_uuid].value
        }
        print(f"\nExcel existente encontrado con {ws.max_row - 1} fila(s). Se agregará lo nuevo.")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas"
        ws.append(ENCABEZADOS)
        for celda in ws[1]:
            celda.font = Font(bold=True)
        uuids_existentes = set()
        print("\nNo existía el Excel, se creará uno nuevo.")

    # ------------------------------------------
    # Procesar cada XML y agregar filas nuevas (evitando duplicados por UUID)
    # ------------------------------------------

    nuevas = 0
    omitidas = 0

    for archivo in archivos_xml:
        print(f"\nLeyendo: {archivo.name}")
        try:
            filas = extraer_filas(archivo)
        except ET.ParseError as e:
            print(f"  ERROR: no se pudo leer el XML ({e}). Se omite.")
            continue

        for fila in filas:
            uuid = fila[-1]
            if uuid in uuids_existentes:
                print(f"  Ya estaba en el Excel (UUID {uuid}), se omite.")
                omitidas += 1
                continue

            ws.append(fila)

            # Formato de moneda para IMPORTE, IVA, TOTAL
            fila_num = ws.max_row
            for col in ("J", "K", "L"):
                ws[f"{col}{fila_num}"].number_format = '$#,##0.00'

            uuids_existentes.add(uuid)
            nuevas += 1

    # ------------------------------------------
    # Ajustar ancho de columnas y guardar
    # ------------------------------------------

    for columna in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in columna
        )
        ws.column_dimensions[columna[0].column_letter].width = min(max_len + 2, 45)

    wb.save(archivo_salida)

    print(f"\n{nuevas} fila(s) nueva(s) agregada(s).")
    if omitidas:
        print(f"{omitidas} fila(s) ya existían y se omitieron.")
    print(f"Excel guardado en: {archivo_salida}")


if __name__ == "__main__":
    main()